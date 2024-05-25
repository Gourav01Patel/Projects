import pandas as pd
from openpyxl import workbook,worksheet,load_workbook
import openpyxl 
from openpyxl.styles import Font, Color, PatternFill, Border, Side,Alignment


Date = input("Please enter Report Date(DDMMYYYY) :-")


Raw_file =r'D:\INTRADAY SHORT ALLOCATION\Input\INTRADAY_EOD_MERGE_'+Date+'.xlsx'
short = r'D:\INTRADAY SHORT ALLOCATION\Input\SHORTALLOCATION_TM_313_'+Date+'.csv'
Output = r'D:\INTRADAY SHORT ALLOCATION\Output\SHORTALLOCATION_TM_313_'+Date+'_Final.xlsx'

shorthead=['TRADE DATE','CLEARING MEMBER CODE','TRADING MEMBER CODE','CLIENT CODE','CP CODE','ALLOCATED COLLATERAL','MARGIN REQUIRED','SHORT ALLOCATION','INDICATOR','REASON CODE',
'EXCESS COLLATERAL NCL','EXCESS COLLATERAL MCCIL','EXCESS COLLATERAL MCXCCL','EXCESS COLLATERAL NCCL','EPI','20 % OF EPI','UCC SUB TYPE','MCX BOD','MCX EOD']


Raw_file1=pd.read_excel(Raw_file,index_col=1,engine='openpyxl')

short1=pd.read_csv(short,header=None,names=shorthead,index_col=1)
short1['EPI'] = short1['CLIENT CODE'].map(Raw_file1.set_index('ctermcode')['EPI'])
short1['20 % OF EPI'] = short1['EPI'] * 20/100
short1['UCC SUB TYPE'] = short1['CLIENT CODE'].map(Raw_file1.set_index('ctermcode')['UCC Cat'])
short1['MCX BOD'] = short1['CLIENT CODE'].map(Raw_file1.set_index('ctermcode')['MCX BOD'])
short1['MCX EOD'] = short1['CLIENT CODE'].map(Raw_file1.set_index('ctermcode')['MCX EOD'])
short1.to_excel(Output,index=False)

wb = openpyxl.load_workbook(Output)
Sheet = wb.active  # Access the active sheet
font_style = Font(name='Book Antiqua', size=9, bold=False, italic=False)
border_style_first_row = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))  # Black thin border
alignment_style_first_row = Alignment(horizontal='center', vertical='center')  # Center alignment
# Iterate through each worksheet in the workbook
for sheet in wb:
    # Iterate through each row in the worksheet
    for row in sheet.iter_rows():
        # Iterate through each cell in the row
        for cell in row:
            # Apply the font style to the cell
            cell.font = font_style
            cell.border= border_style_first_row
            cell.alignment=alignment_style_first_row

wb.save(Output) 

font_style_first_row = Font(name='Book Antiqua',size=9,color='FFFFFF', bold=True)  # Red color and bold
alignment_style_first_row = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Center alignment
fill_style_first_row = PatternFill(start_color="0F2667", end_color="0F2667", fill_type="solid")  # Red background color
# Iterate through each worksheet in the workbook
for sheet in wb:
    # Get the first row
    first_row = sheet[1]

    # Apply font style and alignment to each cell in the first row
    for cell in first_row:
        cell.font = font_style_first_row
        cell.alignment = alignment_style_first_row
        cell.fill = fill_style_first_row
wb.save(Output) 



















print ("Required File Generated to the Path:-",Output)

input("Please Enter any Key for Exit........")

