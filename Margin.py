import cx_Oracle
import time
import os
import pandas as pd
import numpy as np
from datetime import date
import openpyxl 
from openpyxl import workbook,load_workbook
from openpyxl.styles import Font, Color, PatternFill, Border, Side,Alignment
import tkinter as tk
from tkinter import messagebox


Date = input('Please enter report date: ')
#FO_Date = input('Please enter date for FO File (YYYYMMDD): ')
Margin = r'D:\Margin_Shortage\GlobalMarginExport.xlsx'
EQ = r'D:\Margin_Shortage\EQ_MRTM_313.M01'
FO = r'D:\Margin_Shortage\MG0313.M01'
Curr = r'D:\Margin_Shortage\BFX_MGTM0313.M01'
MCX = r'D:\Margin_Shortage\MCX_MARGIN_'+Date[4:] + Date[2:4] +Date[:2]+'_.M01'
NCDEX = r'D:\Margin_Shortage\01274_MGN_' +Date+'.CSV'
Terminal_BR = r'D:\Margin_Shortage\ClientQuery.csv'
MTF = r'D:\Margin_Shortage\mtf_collat_short_margin_'+Date[4:] + Date[2:4] +Date[:2]+'.csv'
output =r'D:\Margin_Shortage\Output\Margin_Shortage_Final'+Date+'.xlsx'
Sort_mtf = r'D:\Margin_Shortage\Output\MTF_Shortage_'+Date+'.xlsx'
test = r'D:\Margin_Shortage\test.csv'

#****************************************************************************************************************************************************************************************************
username = 'ldbo'
password = 'ldbo0313'
host = '192.168.20.74'
port = '1521'
service_name = 'ari2021'

# Construct connection string
dsn = cx_Oracle.makedsn(host, port, service_name=service_name)

# Establish connection
connection = cx_Oracle.connect(username, password, dsn)

# Test connection
cursor = connection.cursor()

# Execute the SQL query
cursor.execute(f"""Select trim(a.oowncode) as CLIENTCODE,trim(a.FIRMNUMBER)as FIRMNUMBER ,trim(a.BRCODE) AS Branch_Code ,trim(b.pangir) as PAN_Number,trim(a.ctermcode) as Terminal_Code
from accounts  a , accountaddressdetail b where a.oowncode=b.oowncode and a.firmnumber='ACML-00001' and b.firmnumber='ACML-00001'  order by a.oowncode""")

# Fetch all rows
rows = cursor.fetchall()

# Convert the fetched data into a pandas DataFrame
df = pd.DataFrame(rows, columns=[desc[0] for desc in cursor.description])

# Export DataFrame to a CSV file
df.to_csv(Terminal_BR, index=False)

# Close cursor and connection
cursor.close()
connection.close()


#************************************************************************************************************************************************************************************************

if os.path.exists(EQ):
    print("Equity File found.")
else:
        # Display a warning pop-up message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showwarning("File Not Found", f"The Equity file is not available at: {EQ}.\nPlease ensure the file is available at the specified path.")

EQ_dtype = {'Date': str, 'Eq Code': str, 'Intial': float, 'Other': float, 'ELM': float, 'Peak': str, 'Adhock': float,
            'Total Margin': float, 'Sym': str, 'Margin required': float,
            'Peak Required': float, 'Margin Collected': float, 'Peak collected': float,
            'Margin Shortage': float, 'Peak Shortage': float, 'EQ Shortage': float}
EQ_head = ['Date', 'Eq Code', 'Intial', 'Other', 'ELM', 'Peak', 'Adhock', 'Total Margin', 'Sym', 'Margin required',
           'Peak Required', 'Margin Collected', 'Peak collected', 'Margin Shortage', 'Peak Shortage', 'EQ Shortage']
# Read Excel file with the provided header and data types
eq1 = pd.read_csv(EQ, header=None, names=EQ_head, dtype=EQ_dtype)
# Calculate Total Shortage
eq1['Margin Shortage'] = eq1['Margin required'] - eq1['Margin Collected']
# Calculate Peak Shortage
eq1['Peak Shortage'] = eq1['Peak Required'] - eq1['Peak collected']
# Calculate Max Shortage
eq1['EQ Shortage'] = eq1[['Margin Shortage', 'Peak Shortage']].max(axis=1)
#*************************************************************************************************************************************************************************************************
#Fo File Process
if os.path.exists(FO):
    print("Future File found.")
else:
        # Display a warning pop-up message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showwarning("File Not Found", f"The Future file is not available at: {FO}.\nPlease ensure the file is available at the specified path.")

def progress_bar(total_iterations):
    for i in range(total_iterations + 1):
        progress = (i / total_iterations) * 100
        print(f"\rProcessing...{progress:.2f}%", end='', flush=True)
        time.sleep(0.10)  # Simulate some work being done
    print("Process Complete")



FO_dtype = {'Date':str,'Fo Code':str,'Intital':float,'Excposre':float,'E':float,'Other':float,'G':float,
            'H':float,'Peak':float,'Total Margin':float,'K':str,'Total Collected':float,
            'Peak Collected':float,'Total Shortage':float,'Peak shortage':float,'FO Shortage':float}
FO_head = ['Date','Fo Code','Intital','Excposre','E','Other','G','H','Peak','Total Margin',
           'K','Total Collected','Peak Collected','Total Shortage','Peak shortage','FO Shortage']
# Read Excel file with the provided header and data types
fo1 = pd.read_csv(FO, header=None, names=FO_head, dtype=FO_dtype)
# Calculate Total Shortage
fo1['Total Shortage'] = fo1['Total Margin'] - fo1['Total Collected']
# Calculate Peak Shortage
fo1['Peak Shortage'] = fo1['Peak'] - fo1['Peak Collected']
# Calculate Max Shortage
fo1['FO Shortage'] = fo1[['Total Shortage', 'Peak Shortage']].max(axis=1)

#*************************************************************************************************************************************************************************************************
#Currency File Process

if os.path.exists(Curr):
    print("Currency File found.")
else:
        # Display a warning pop-up message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showwarning("File Not Found", f"The Currency file is not available at: {Curr}.\nPlease ensure the file is available at the specified path.")


Curr_dtype = {'Date':str,'Curr Code':str,'Intital':float,'Excposre':float,'Other':float,'MTM':float,'G':float,'Peak':float,
              'Total Margin':float,'J':str,'Total Collected':float,'Peak Collected':float,
              'Peak Shortage':float,'Margin Shortage':float,'Max Shortage':float}
Curr_head = ['Date','Curr Code','Intital','Excposre','Other','MTM','G','Peak','Total Margin','J','Total Collected',
             'Peak Collected','Peak Shortage','Margin Shortage','Curr Shortage']
# Read Excel file with the provided header and data types
Curr1 = pd.read_csv(Curr, header=None, names=Curr_head, dtype=Curr_dtype)
# Calculate Total Shortage
Curr1['Margin Shortage'] = Curr1['Total Margin'] - Curr1['Total Collected']
# Calculate Peak Shortage
Curr1['Peak Shortage'] = Curr1['Peak'] - Curr1['Peak Collected']
# Calculate Max Shortage
Curr1['Curr Shortage'] = Curr1[['Margin Shortage', 'Peak Shortage']].max(axis=1)

#**************************************************************************************************************************************************************************************************
#MCX File Process

if os.path.exists(MCX):
    print("MCX File found.")
else:
        # Display a warning pop-up message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showwarning("File Not Found", f"The MCX file is not available at: {MCX}.\nPlease ensure the file is available at the specified path.")



MCX_dtype = {'Date':str,'Member':str,'MCX Code':str,'Initial':float,'Other':float,'M2M':float,'G':str,'H':str,'M2M Collected':float,
             'Initial Collected':float,'Other Collected':float,'L':float,'Peak':float,'Peak Collected':float,'O':str,
             'P':float,'Q':float,'MTM Shortage':float,'Initial Shortage':float,'Other Shortage':float,'Peak Shortage':float,'MCX Max':float,
             'MCX Shortage':float}
MCX_head = ['Date','Member','MCX Code','Initial','Other','M2M','G','H','M2M Collected','Initial Collected',
            'Other Collected','L','Peak','Peak Collected','O','P','Q','MTM Shortage','Initial Shortage',
            'Other Shortage','Peak Shortage','MCX Max','MCX Shortage']
# Read Excel file with the provided header and data types
MCX1 = pd.read_csv(MCX, header=None, names=MCX_head, dtype=MCX_dtype,na_values=[' ', 'NA', 'N/A'])

#MCX1 = pd.read_csv(MCX, header=None, names=MCX_head, dtype=MCX_dtype, na_values=[' ', 'NA', 'N/A'])


# Calculate MTM Shortage
MCX1['MTM Shortage'] = MCX1['M2M'] - MCX1['M2M Collected']
# Calculate Initial Shortage
MCX1['Initial Shortage'] = MCX1['Initial'] - MCX1['Initial Collected']
# Calculate Other Shortage
MCX1['Other Shortage'] = MCX1['Other'] - MCX1['Other Collected']
# Calculate Peak Shortage
MCX1['Peak Shortage'] = MCX1['Peak'] - MCX1['Peak Collected']

# Calculate Max Shortage
MCX1['MCX Max'] = MCX1[['Initial Shortage','Peak Shortage']].max(axis=1)

MCX1['MCX Shortage'] = MCX1[['MCX Max','MTM Shortage','Other Shortage']].sum(axis=1)

#***************************************************************************************************************************************************************************************************

#NCDEX File Process
if os.path.exists(NCDEX):
    print("NCDEX File found.")
else:
        # Display a warning pop-up message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showwarning("File Not Found", f"The NCDEX file is not available at: {NCDEX}.\nPlease ensure the file is available at the specified path.")



NCDEX_dtype = {'Trade Date':str,	'Trading Member ID':str,	'Client ID':str,'Initial Mrgin (Initial + Extreme Loss)':float,	'Peak Initial Mrgin':float,	'Other Mrgns':float,
               	'MTM (Gain/Loss)':float,'Upfront initial Mrgn Collected':float,	'Peak Initial Mrgn Collected':float,	'Other Margins Collected by T+2':float,
               	'MTM (Gains)/ Loss Collected By T+2':float,	'Shortfall in Initial Mrgn':float,	'Shortfall in Peak Mrgn':float,
               	'Shortfall in Other Mrgn':float,'Shortfall in MTM (Gain)/Loss':float,'Initial Margin at Peak Short Allocation (IM-PSA)':float,
               	'Collateral Value at EOD':float,'Collateral Value at Peak Short Allocation (PSA)':float,
               	'Excess Collateral With Other CCs (ICCL)':float,'Excess Collateral With Other CCs (MCXCCL)':float,
               	'Excess Collateral With Other CCs (NCL)':float,'Excess Collateral With Other CCs (Others)':float,
               	'Delay In Processing Allocation Request':float,	'Trade In Wrong Client Code':float,	'Total Shortfall (EOD Short Allocation)':float,
               	'Total Shortfall (Peak Short Allocation)':float,'Initial Shortage':float,'Peak Shortage':float,'Other Shortage':float,
                'MTM Shortage':float,'NCDEX Max':float,'NCDEX Shortage':float}
NCDEX_head = ['Trade Date',	'Trading Member ID','Client ID','Initial Mrgin (Initial + Extreme Loss)',
            'Peak Initial Mrgin','Other Mrgns','MTM (Gain/Loss)','Upfront initial Mrgn Collected','Peak Initial Mrgn Collected',
            'Other Margins Collected by T+2','MTM (Gains)/ Loss Collected By T+2','Shortfall in Initial Mrgn',
            'Shortfall in Peak Mrgn','Shortfall in Other Mrgn','Shortfall in MTM (Gain)/Loss',
            'Initial Margin at Peak Short Allocation (IM-PSA)','Collateral Value at EOD','Collateral Value at Peak Short Allocation (PSA)',
            'Excess Collateral With Other CCs (ICCL)','Excess Collateral With Other CCs (MCXCCL)',
            'Excess Collateral With Other CCs (NCL)','Excess Collateral With Other CCs (Others)',
            'Delay In Processing Allocation Request','Trade In Wrong Client Code','Total Shortfall (EOD Short Allocation)',
            'Total Shortfall (Peak Short Allocation)','Initial Shortage','Peak Shortage','Other Shortage ','MTM Shortage','NCDEX Max','NCDEX Shortage']
# Read Excel file with the provided header and data types
NCDEX1 = pd.read_csv(NCDEX, header=None, names=NCDEX_head, dtype=NCDEX_dtype,na_values=[' ', 'NA', 'N/A'], skiprows=1)
#NCDEX1 = NCDEX1.replace("'", "")
NCDEX1['Client ID'] = NCDEX1['Client ID'].str.replace("'", '')
# Calculate Total Shortage
NCDEX1['Initial Shortage'] = NCDEX1['Initial Mrgin (Initial + Extreme Loss)'] - NCDEX1['Upfront initial Mrgn Collected']
# Calculate Peak Shortage
NCDEX1['Peak Shortage'] = NCDEX1['Peak Initial Mrgin'] - NCDEX1['Peak Initial Mrgn Collected']
# Calculate Other Shortage
NCDEX1['Other Shortage'] = NCDEX1['Other Mrgns'] - NCDEX1['Other Margins Collected by T+2']
# Calculate MTM Shortage
NCDEX1['MTM Shortage'] = np.where(NCDEX1['MTM (Gain/Loss)'] < 0,0,
                                  NCDEX1['MTM (Gain/Loss)']) - NCDEX1['MTM (Gains)/ Loss Collected By T+2']
# Calculate Max Shortage   G2 = MTM (Gain/Loss)    MTM (Gains)/ Loss Collected By T+2  = K2

NCDEX1['NCDEX Max'] = NCDEX1[['Initial Shortage','Peak Shortage']].max(axis=1)

NCDEX1['NCDEX Shortage']= NCDEX1[['NCDEX Max','Other Shortage','MTM Shortage']].sum(axis=1)

#************************************************************************************************************************************************************************************************

Terminal_BR_Head = ['CLIENTCODE','FIRMNUMBER','Branch_Code','PAN_Number','Terminal_Code']
Terminal_BR_Dtype = {'CLIENTCODE':str,'FIRMNUMBER':str,'Branch_Code':str,'PAN_Number':str,'Terminal_Code':str}

Terminal_BR1 = pd.read_csv(Terminal_BR, header=None, names=Terminal_BR_Head, dtype=Terminal_BR_Dtype,na_values=[' ', 'NA', 'N/A'], skiprows=1)


if os.path.exists(Terminal_BR):
    print("Terminal And Branch File found.")
else:
        # Display a warning pop-up message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showwarning("File Not Found", f"The Terminal And Branch file is not available at: {Terminal_BR}.\nPlease ensure the file is available at the specified path.")

#*********************************************************************************************************************************************************************************************

if os.path.exists(Margin):
    print("LDFILE File found.")
else:
        # Display a warning pop-up message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showwarning("File Not Found", f"The LDFILE file is not available at: {Margin}.\nPlease ensure the file is available at the specified path.")

Margin_Head = ['Client Code',	'Client Name',	'Finance Dr-Cr',	'MTM Credit',	'MTM 2 Plus Credit',	'Early Payin',
            'EP Additional Credit',	'SEBI MTF Credit',	'Peak Payment credit','Cash Peak Payment Credit',	'Peak Stock credit',
            'Margin  Dr-Cr','Premium Cr',	'Span Margin','Collateral - I','Collateral - II',	'BenToPool Stock','CUSPA Stock Val','Snap 1 Peak Margin',
            'Snap Min Credit','T1 Del Credit','POA Stocks',	'Free POA Stocks',	'Demat Stocks',	'Final  Dr-Cr',	'Exchange Total Margin',
            'Terminal Code','Branch','FO Shortage','EQ Shortage','Curr Shortage','MCX Shortage','NCDEX Shortage','Total Shortage','Final Shortage']
Margin_Dtype = {'Client Code':str,'Client Name':str,'Finance Dr-Cr':float,'MTM Credit':float,'MTM 2 Plus Credit':float,
'Early Payin':float,'EP Additional Credit':float,'SEBI MTF Credit':float,'Peak Payment credit':float,'Cash Peak Payment Credit':float,
'Peak Stock credit':float,'Margin  Dr-Cr':float,'Premium Cr':float,'Span Margin':float,'Collateral - I':float,'Collateral - II':float,
'BenToPool Stock':float,'CUSPA Stock Val':float,'Snap 1 Peak Margin':float,'Snap Min Credit':float,'T1 Del Credit':float,'POA Stocks':float,
'Free POA Stocks':float,'Demat Stocks':float,'Final  Dr-Cr':float,'Exchange Total Margin':float,'Terminal Code':str,'Branch':str,
'FO Shortage':float,'EQ Shortage':float,'Curr Shortage':float,'MCX Shortage':float,'NCDEX Shortage':float,'Total Shortage':float,'Final Shortage':float}

sheet='Margin Breakup'
sheet_S='Margin Status'

Margin1 = pd.read_excel(Margin, header=None, names=Margin_Head, dtype=Margin_Dtype,na_values=[' ', 'NA', 'N/A'],sheet_name=sheet, skiprows=5)
#Margin1.ffill(None,'0',inplace=True)
Margin1['Terminal Code'] = Margin1['Client Code'].map(Terminal_BR1.set_index('CLIENTCODE')['Terminal_Code'])
Margin1['Terminal Code']=Margin1['Terminal Code'].fillna(0)
Margin1['Branch'] = Margin1['Client Code'].map(Terminal_BR1.set_index('CLIENTCODE')['Branch_Code'])
Margin1['Branch']=Margin1['Branch'].fillna(0)
Margin1['EQ Shortage'] = Margin1['Terminal Code'].map(eq1.set_index('Eq Code')['EQ Shortage'])
Margin1['EQ Shortage']=Margin1['EQ Shortage'].fillna(0)
Margin1['FO Shortage'] = Margin1['Terminal Code'].map(fo1.set_index('Fo Code')['FO Shortage'])
Margin1['FO Shortage']=Margin1['FO Shortage'].fillna(0)
Margin1['Curr Shortage'] = Margin1['Terminal Code'].map(Curr1.set_index('Curr Code')['Curr Shortage'])
Margin1['Curr Shortage']=Margin1['Curr Shortage'].fillna(0)
Margin1['MCX Shortage'] = Margin1['Terminal Code'].map(MCX1.set_index('MCX Code')['MCX Shortage'])
Margin1['MCX Shortage']=Margin1['MCX Shortage'].fillna(0)
Margin1['NCDEX Shortage'] = Margin1['Terminal Code'].map(NCDEX1.set_index('Client ID')['NCDEX Shortage'])
Margin1['NCDEX Shortage']=Margin1['NCDEX Shortage'].fillna(0)
Margin1['Total Shortage'] = Margin1[['FO Shortage', 'EQ Shortage', 'Curr Shortage', 'MCX Shortage', 'NCDEX Shortage']].sum(axis=1)

# Apply the logic to calculate Final Shortage

Margin1['Final Shortage'] = np.abs(np.where(Margin1['Final  Dr-Cr'] < 0,(Margin1['Final  Dr-Cr'] - Margin1['Total Shortage']),Margin1['Total Shortage']))


# Assuming 'Final Dr-Cr' and 'Total Shortage' are already existing columns in Margin1 DataFrame
#Filter for Shortage Only > 0

Margin1 = Margin1.loc[Margin1['Total Shortage'] != 0]
Margin1 = Margin1.sort_values(by='Total Shortage', ascending=True)
Margin1.to_excel(output,sheet_name='Summary',index=False)

#**MTF Shortage Working **************************************************************


Margin2_Head = ['Client Code',	'Client Name',	'Finance Dr-Cr',	'MTM Credit',	'MTM 2 Plus Credit',	'Early Payin',
            'EP Additional Credit',	'SEBI MTF Credit',	'Peak Payment credit','Cash Peak Payment Credit',	'Peak Stock credit',
            'Margin  Dr-Cr','Premium Cr',	'Span Margin','Collateral - I','Collateral - II',	'BenToPool Stock','CUSPA Stock Val',	'Snap 1 Peak Margin',
            'Snap Min Credit','T1 Del Credit','POA Stocks',	'Free POA Stocks',	'Demat Stocks',	'Final  Dr-Cr',	'Exchange Total Margin','MTF Collateral','MTF Shortage']
Margin2_Dtype = {'Client Code':str,'Client Name':str,'Finance Dr-Cr':float,'MTM Credit':float,'MTM 2 Plus Credit':float,
'Early Payin':float,'EP Additional Credit':float,'SEBI MTF Credit':float,'Peak Payment credit':float,'Cash Peak Payment Credit':float,
'Peak Stock credit':float,'Margin  Dr-Cr':float,'Premium Cr':float,'Span Margin':float,'Collateral - I':float,'Collateral - II':float,
'BenToPool Stock':float,'CUSPA Stock Val':float,'Snap 1 Peak Margin':float,'Snap Min Credit':float,'T1 Del Credit':float,'POA Stocks':float,
'Free POA Stocks':float,'Demat Stocks':float,'Final  Dr-Cr':float,'Exchange Total Margin':float,'MTF Collateral':float,'MTF Shortage':float}

sheet2='Margin Breakup'

Margin2 = pd.read_excel(Margin, header=None, names=Margin2_Head, dtype=Margin2_Dtype,na_values=[' ', 'NA', 'N/A'],sheet_name=sheet2, skiprows=5,index_col=1)

#Margin3 = pd.read_excel(Margin, header=None, names=Margin2_Head,dtype=Margin2_Dtype,na_values=[' ', 'NA', 'N/A'],sheet_name=sheet2, skiprows=5,index_col=1


MTF_Head = ['oowncode','mtf_collateral']
MTF_dtype = {'oowncode':str,'mtf_collateral':float}

MTF1=pd.read_csv(MTF,names=MTF_Head,dtype=MTF_dtype,skiprows=1)

MTF1.replace('"','',inplace=True)

Margin2['MTF Collateral'] = Margin2['Client Code'].map(MTF1.set_index('oowncode')['mtf_collateral'])

#Margin2['MTF Collateral'].to_csv(test,index=False)


Margin2['MTF Shortage'] = np.round(np.where(
    Margin2['Exchange Total Margin'] > 0,
    (Margin2['Final  Dr-Cr'] - Margin2['Exchange Total Margin'] - Margin2['MTF Collateral']),
    4  # This is the default value if the condition is not met, adjust as needed
))

Margin2 = Margin2.loc[Margin2['MTF Shortage'] < 0]

Margin2.to_excel(Sort_mtf,sheet_name='MTF_Shortage',index=False)


#Formatting Part#*********************************************************************

wb = openpyxl.load_workbook(output)
sheet = wb.active  # Access the active sheet



font_style = Font(name='Book Antiqua', size=9, bold=False, italic=False)
border_style_first_row = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))  # Black thin border

# Iterate through each worksheet in the workbook
for sheet in wb:
    # Iterate through each row in the worksheet
    for row in sheet.iter_rows():
        # Iterate through each cell in the row
        for cell in row:
            # Apply the font style to the cell
            cell.font = font_style
            cell.border= border_style_first_row

wb.save(output) 

font_style_first_row = Font(name='Book Antiqua',size=10,color='FFFFFF', bold=True)  # Red color and bold
alignment_style_first_row = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Center alignment
fill_style_first_row = PatternFill(start_color="0F2667", end_color="0F2667", fill_type="solid")  # Red background color
# Iterate through each worksheet in the workbook
for sheet in wb:
    # Get the first row
    first_row = sheet[1]

    # Apply font style and alignment to each cell in the first row
    for cell in first_row:
        cell.font = font_style_first_row
        cell.alignment = alignment_style_first_row
        cell.fill = fill_style_first_row
wb.save(output) 

##MTF File Formating************************************************

wb = openpyxl.load_workbook(Sort_mtf)
sheet = wb.active  # Access the active sheet



font_style = Font(name='Book Antiqua', size=9, bold=False, italic=False)
border_style_first_row = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))  # Black thin border

# Iterate through each worksheet in the workbook
for sheet in wb:
    # Iterate through each row in the worksheet
    for row in sheet.iter_rows():
        # Iterate through each cell in the row
        for cell in row:
            # Apply the font style to the cell
            cell.font = font_style
            cell.border= border_style_first_row

wb.save(Sort_mtf) 
font_style_first_row = Font(name='Book Antiqua',size=10,color='FFFFFF', bold=True)  # Red color and bold
alignment_style_first_row = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Center alignment
fill_style_first_row = PatternFill(start_color="0F2667", end_color="0F2667", fill_type="solid")  # Red background color
# Iterate through each worksheet in the workbook
for sheet in wb:
    # Get the first row
    first_row = sheet[1]

    # Apply font style and alignment to each cell in the first row
    for cell in first_row:
        cell.font = font_style_first_row
        cell.alignment = alignment_style_first_row
        cell.fill = fill_style_first_row
wb.save(Sort_mtf) 

#Formatting End#*****************************************************

total_iterations = 60
  # Total number of iterations in your program
progress_bar(total_iterations)


print("Report is Generated to the path:-",output)
input("Press Enter to exit...")




